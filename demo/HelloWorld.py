#author: yangxuanlun  Email:yangxuanlun@163.com time:2019/11/10
# from selenium import webdriver
# from bs4 import BeautifulSoup
# import re
# import time  http://2211.xyz/?m=vod-play-id-25831-src-1-num-1.html

import time
from selenium import webdriver
# from selenium.webdriver.common.keys import Keys
import os
# # import win32con
# # import win32api
import pyautogui
import psutil
import openpyxl

class mTestD2c():
    def __init__(self):
        self.mUnit = 500
        self.mTo = 1600
        self.pyGui = pyautogui
        self.row = 1

    def move_founction(self, mTag, worksheet,ncols):
        mTD = self
        for i in range(10):
            #print(i)
            if mTag == 0:
                pyautogui.dragTo(mTD.mTo, mTD.mUnit, duration=0.2)
                pyautogui.moveTo(mTD.mUnit, mTD.mUnit, duration=0)
            else:
                pyautogui.dragTo(mTD.mUnit, mTD.mUnit, duration=0.2)
                pyautogui.moveTo(mTD.mTo, mTD.mUnit, duration=0)
            cpu = (psutil.cpu_percent(1))
            memory = psutil.Process(os.getpid()).memory_info().rss
            #worksheet.write(mTD.row, ncols+1, cpu)
            mTD.row = mTD.row + 1
            worksheet.cell(row=mTD.row, column=ncols+1).value = cpu
            worksheet.cell(row=mTD.row, column=ncols + 2).value = memory


if __name__ == '__main__':

    # url = 'http://192.168.11.160/compare_with_d2c_and_mapbox/d2c.html'
    print('请输入url：')
    url = input()
    print('请输入测试名称')
    name = input()
    workbook = openpyxl.load_workbook('test1.xlsx')
    worksheet = workbook.worksheets[0]
    ncols = worksheet.max_column
    worksheet.cell(row=1, column=ncols+1).value = name

    options = webdriver.ChromeOptions()
   # options.add_argument('user-agent="Mozilla/5.0 (Linux; Android 4.0.4; Galaxy Nexus Build/IMM76B) AppleWebKit/535.19 (KHTML, like Gecko) Chrome/18.0.1025.133 Mobile Safari/535.19"')
    driver = webdriver.Chrome(chrome_options=options)
    driver.get(url)
    html = driver.page_source
    time.sleep(8)
    mTD = mTestD2c()

    pyautogui.moveTo(mTD.mUnit, mTD.mUnit, duration=0)
    for m in range(10):
        mTD.move_founction(m % 2, worksheet, ncols)
        if m<6:
            pyautogui.scroll(-1000)
            pyautogui.scroll(-1000)
            time.sleep(0.2)
        else:
            pyautogui.scroll(1000)
            time.sleep(0.2)
    workbook.save(filename='test1.xlsx')










