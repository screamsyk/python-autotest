import openpyxl


def main():  # 官方说明：https://openpyxl.readthedocs.io

    # 工作簿与工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    wb.create_sheet('sheet1')
    print(wb.sheetnames)
    ws1 = wb['sheet1']
    ws1.sheet_properties.tabColor = 'eeffcc'

    # 操作单元格
    c = ws['A1']
    c.value = '测试 openpyxl 的功能'
    c.font = openpyxl.styles.Font(bold=True)
    ws.row_dimensions[1].height = 30
    ws.cell(row=1, column=2, value='操作单元格')

    # 合并单元格
    ws.merge_cells('A2:D2')
    ws.merge_cells(start_row=3, start_column=1, end_row=6, end_column=1)

    # 保存文件
    wb.save('test_openpyxl.xlsx')


if __name__ == '__main__':
    main()
