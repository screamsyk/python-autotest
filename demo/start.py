# !/usr/bin/env python3
# -*- coding: utf-8 -*-

# 内置模块
import os
import sys
import json
import threading
import time
from datetime import datetime

# 第三方模块
from selenium import webdriver
import psutil
import openpyxl

# 自定义模块
from config import *

driver = webdriver.Chrome()  # 浏览器驱动
is_running = False
records = []


def map_worker():  # (1)模拟地图操作
    global driver
    global is_running

    # 地图缩放
    driver.execute_script(
        'd2cMap.setCenter([106.57577176873292,29.565013976572004])')
    for zoom in zooms:
        driver.execute_script('d2cMap.setZoom(' + str(zoom) + ')')
        time.sleep(1)

    # 地图移动
    driver.execute_script('d2cMap.setZoom(16)')
    for point in points:
        driver.execute_script('d2cMap.setCenter(' + (str)(point) + ')')
        time.sleep(1)

    # 结束地图操作
    is_running = False


def record_worker():  # (2)记录 CPU 和内存等信息
    global records
    while is_running:
        time = datetime.now().strftime('%H:%M:%S')
        cpu_percent = psutil.cpu_percent(1)
        memory_percent = psutil.virtual_memory().percent
        memory = psutil.virtual_memory().used
        record = {'time': time, 'cpu_percent': cpu_percent, 'memory': memory,
                  'memory_percent': memory_percent}
        records.append(record)
    save()
    stat()


def save():  # (3)保存信息为 json 文件
    global driver
    global records
    file_path = 'result_'+driver.title+'.json'
    obj = {'title': driver.title, 'records': records}
    with open(file_path, 'w+') as f:
        f.write(json.dumps(obj))
    pass


def stat():  # (4)统计，并记录到 excel 中
    global records
    global driver
    cpu_percent_sum = 0
    cpu_percent_avg = 0
    cpu_percent_over = 0
    memory_sum = 0
    memory_avg = 0
    memory_percent_sum = 0
    memory_percent_avg = 0
    memory_percent_over = 0
    workbook = openpyxl.load_workbook('result.xlsx')
    worksheet = workbook.worksheets[0]
    ncols = worksheet.max_column
    worksheet.cell(row=1, column=ncols+1).value = driver.title
    worksheet.cell(row=2, column=ncols+1).value = 'CUP 占用率'
    worksheet.cell(row=2, column=ncols+2).value = '内存占用大小'
    worksheet.cell(row=2, column=ncols+3).value = '内存占用率'
    row = 3
    for record in records:
        cpu_percent_sum = cpu_percent_sum+record['cpu_percent']
        memory_sum = memory_sum+record['memory']
        memory_percent_sum = memory_percent_sum+record['memory_percent']
        if record['cpu_percent'] >= 90:
            cpu_percent_over = cpu_percent_over+1
        if record['memory_percent'] >= 90:
            memory_percent_over = memory_percent_over+1
        worksheet.cell(row=row, column=ncols + 1).value = record['cpu_percent']
        worksheet.cell(row=row, column=ncols + 2).value = record['memory']
        worksheet.cell(row=row, column=ncols +
                       3).value = record['memory_percent']
        row = row+1
    workbook.save(filename='result.xlsx')
    cpu_percent_avg = cpu_percent_sum/len(records)
    memory_avg = memory_sum/len(records)
    memory_percent_avg = memory_percent_sum/len(records)
    print('CPU 平均占用率：', float('%.2f' % cpu_percent_avg), '%')
    print('CPU 占用率达 90% 及以上次数：', cpu_percent_over)
    print('内存平均占用大小：', int(memory_avg))
    print('内存平均占用率：', float('%.2f' % memory_percent_avg), '%')
    print('内存占用率达 90% 及以上次数：', memory_percent_over)


def start():  # (5)开始执行
    global driver
    global is_running
    global records
    driver.maximize_window()
    driver.get(url)
    is_running = True
    records = []
    time.sleep(10)
    map_thread = threading.Thread(target=map_worker, name='map_worker')
    record_thread = threading.Thread(
        target=record_worker, name='record_worker')
    map_thread.start()
    record_thread.start()


start()
