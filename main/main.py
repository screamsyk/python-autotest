# 内置模块
from datetime import datetime
from time import sleep
from threading import Thread
import json
import os

# 第三方模块
import configparser
import psutil
import openpyxl
from selenium import webdriver
import pyautogui
from pyecharts.charts import Line
from pyecharts.globals import ThemeType
from pyecharts import options as opts
import keyboard

# 自定义模块
import mock


# 基本数据
g_mode = 'autogui'
g_test_name = '测试'
g_excel_name = 'result'
g_url = ''
g_map_type = 'd2c'
g_run_minutes = 0
g_driver = None
g_is_running = False
g_records = []
g_start_time = None
g_end_time = None


def read_config():  # (1)读取配置文件 config.ini
    global g_mode
    global g_test_name
    global g_excel_name
    global g_url
    global g_run_minutes
    global g_map_type
    config = configparser.ConfigParser()
    config.read('config.ini', encoding="utf-8-sig")  # 编码为 utf-8-sig ，避免中文乱码
    if(not config.sections()):
        print('无配置文件 config.ini')
        return
    g_mode = config['params']['mode']
    g_test_name = config['params']['test_name']
    g_excel_name = config['params']['excel_name']
    g_url = config['url'][config['params']['url_name']]
    g_run_minutes = int(config['params']['run_minutes'])
    g_map_type = config['params']['map_type']


def test_by_javascript():  # (2)用 javascript 控制地图移动和缩放
    global g_is_running

    while not is_stop():

        # 地图还原
        if g_map_type == 'd2c':
            g_driver.execute_script(
                f'd2cMap.setCenter({mock.map_view["center"]})')
            g_driver.execute_script(f'd2cMap.setZoom({mock.map_view["zoom"]})')
        elif g_map_type == 'baidu':
            g_driver.execute_script(
                f'map.setCenter(new BMap.Point({mock.baidu_map_view["lng"]},{mock.baidu_map_view["lat"]}))')
            g_driver.execute_script(f'map.setZoom({mock.map_view["zoom"]})')

        # 地图缩放
        for zoom in mock.zooms:
            if is_stop():
                break
            if g_map_type == 'd2c':  # d2c地图
                g_driver.execute_script(f'd2cMap.setZoom({zoom})')
            elif g_map_type == 'baidu':  # 百度地图
                g_driver.execute_script(f'map.setZoom({zoom})')
            sleep(1)

        # 地图移动
        if g_map_type == 'd2c':  # d2c地图
            g_driver.execute_script(f'd2cMap.setZoom(16)')
        elif g_map_type == 'baidu':  # 百度地图
            g_driver.execute_script(f'map.setZoom(16)')
        for point in mock.points:
            if is_stop():
                break
            if g_map_type == 'd2c':  # d2c地图
                g_driver.execute_script(f'd2cMap.setCenter({point})')
            sleep(1)

    # 退出
    g_driver.close()
    g_driver.quit()
    g_is_running = False


def test_by_autogui():  # (3)用 autogui 调用鼠标控制地图移动和缩放
    global g_is_running
    screen_width, screen_height = pyautogui.size()  # 屏幕大小

    while not is_stop():

        # 地图还原
        if g_map_type == 'd2c':  # d2c地图
            g_driver.execute_script(
                f'map.setCenter({mock.map_view["center"]})')
            g_driver.execute_script(f'map.setZoom({mock.map_view["zoom"]})')
        elif g_map_type == 'baidu':  # 百度地图
            g_driver.execute_script(
                f'map.setCenter(new BMap.Point({mock.baidu_map_view["lng"]},{mock.baidu_map_view["lat"]}))')
            g_driver.execute_script(f'map.setZoom({mock.map_view["zoom"]})')
        sleep(2)

        # 鼠标滚动
        pyautogui.moveTo(screen_width / 2, screen_height / 2)  # 鼠标移动
        for scroll in mock.scrolls:
            if is_stop():
                break
            pyautogui.scroll(scroll)  # 鼠标滚轮滚动
            sleep(1.5)

        # 鼠标拖动
        for drag in mock.drags:
            if is_stop():
                break
            pyautogui.moveTo(drag['from'][0], drag['from'][1])
            pyautogui.dragTo(drag['to'][0], drag['to'][1], duration=0.2)
            sleep(1.5)

    # 退出
    g_driver.close()
    g_driver.quit()
    g_is_running = False


def record_data():  # (4)记录 CPU 、内存等数据
    while g_is_running:
        time = datetime.now().strftime('%H:%M:%S')
        cpu_percent = psutil.cpu_percent(1)  # 每秒统计下 CPU 使用率、内存使用率、已用内存
        memory_percent = psutil.virtual_memory().percent
        memory = psutil.virtual_memory().used
        record = {'time': time, 'cpu_percent': cpu_percent,
                  'memory_percent': memory_percent, 'memory': memory}
        g_records.append(record)
    save_to_excel()
    save_to_json()
    save_to_chart()


def save_to_excel():  # (5)保存数据到 excel
    filename = g_excel_name+'.xlsx'
    if os.path.isfile(filename):
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.worksheets[0]
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
    now_col = worksheet.max_column+1
    worksheet.cell(row=1, column=now_col+1).value = g_test_name
    worksheet.cell(row=3, column=now_col+1).value = 'CUP 使用率'
    worksheet.cell(row=3, column=now_col+2).value = '内存使用率'
    worksheet.cell(row=3, column=now_col+3).value = '已用内存'
    now_row = 4
    cpu_percent_sum = 0
    cpu_percent_avg = 0
    cpu_percent_over = 0
    memory_sum = 0
    memory_avg = 0
    memory_percent_sum = 0
    memory_percent_avg = 0
    memory_percent_over = 0
    for record in g_records:
        cpu_percent_sum = cpu_percent_sum+record['cpu_percent']
        memory_sum = memory_sum+record['memory']
        memory_percent_sum = memory_percent_sum+record['memory_percent']
        if record['cpu_percent'] >= 90:
            cpu_percent_over = cpu_percent_over+1
        if record['memory_percent'] >= 90:
            memory_percent_over = memory_percent_over+1
        worksheet.cell(row=now_row, column=now_col +
                       1).value = record['cpu_percent']
        worksheet.cell(row=now_row, column=now_col +
                       2).value = record['memory_percent']
        worksheet.cell(row=now_row, column=now_col +
                       3).value = record['memory']
        now_row = now_row+1
    length = len(g_records)
    cpu_percent_avg = cpu_percent_sum/length
    memory_avg = memory_sum/length
    memory_percent_avg = memory_percent_sum/length
    worksheet.cell(row=2, column=now_col +
                   1).value = f'统计（{datetime.fromtimestamp(g_start_time)}至{datetime.fromtimestamp(g_end_time)}）：（1）CPU 平均使用率：{float("%.2f" % cpu_percent_avg)}%；（2）CPU 使用率达 90% 及以上次数：{cpu_percent_over}；（3）内存平均使用率：{float("%.2f" % memory_percent_avg)}%；（4）内存使用用率达 90% 及以上次数：{memory_percent_over}；（5）内存平均使用大小：{int(memory_avg)}'
    workbook.save(filename=filename)
    print('测试开始时间：', datetime.fromtimestamp(g_start_time))
    print('测试结束时间：', datetime.fromtimestamp(g_end_time))
    print('CPU 平均使用率：', float('%.2f' % cpu_percent_avg), '%')
    print('CPU 使用率达 90% 及以上次数：', cpu_percent_over)
    print('内存平均使用率：', float('%.2f' % memory_percent_avg), '%')
    print('内存使用用率达 90% 及以上次数：', memory_percent_over)
    print('内存平均使用大小：', int(memory_avg))


def save_to_chart():  # (6)保存到图表
    xaxis = []
    yaxis_cpu_percent = []
    yaxis_memory_percent = []
    for item in g_records:
        xaxis.append(item['time'])
        yaxis_cpu_percent.append(item['cpu_percent'])
        yaxis_memory_percent.append(item['memory_percent'])
    line = (
        Line(
            init_opts={
                'theme': ThemeType.LIGHT,
                'width': '1400px',
                'height': '600px',
                'page_title': g_test_name
            }
        )
        .add_xaxis(xaxis)
        .add_yaxis('CPU 使用率', yaxis_cpu_percent, markline_opts=opts.MarkLineOpts(data=[opts.MarkLineItem(type_="average")]))
        .add_yaxis('内存使用率', yaxis_memory_percent, markline_opts=opts.MarkLineOpts(data=[opts.MarkLineItem(type_="average")]))
        .set_global_opts(
            title_opts={
                'text': g_test_name,
                'pos_left': 'center'
            },
            datazoom_opts={
                'is_show': True
            },
            tooltip_opts={
                'trigger': 'axis'
            },
            xaxis_opts={
                'name': '时间'
            },
            yaxis_opts={
                'name': '百分比',
                'max': 100
            }
        )
    )
    line.render(g_test_name+'.html')


def save_to_json():  # (7)保存数据到 json 文件
    filename = g_test_name+'.json'
    obj = {'title': g_test_name, 'records': g_records}
    with open(filename, 'w+') as f:
        f.write(json.dumps(obj))


def keyboard_listener():  # (8)监听键盘 ctrl+alt+s ，停止程序
    def stop_python():
        global g_is_running
        g_is_running = False
        print('------程序停止------')
    keyboard.add_hotkey('ctrl+alt+s', stop_python)


def is_stop():  # (9)是否停止程序
    global g_end_time
    g_end_time = int(datetime.now().timestamp())
    return (not g_is_running) or (g_end_time-g_start_time > g_run_minutes*60)


# 主程序
if __name__ == '__main__':

    # 读取配置
    read_config()
    print('------配置参数------')
    print(g_mode)
    print(g_test_name)
    print(g_excel_name)
    print(g_url)
    print(f'{g_run_minutes}分钟')

    # 浏览器驱动
    print('------正在打开地图...------')
    g_driver = webdriver.Chrome()
    g_driver.maximize_window()
    g_driver.get(g_url)
    sleep(10)

    # 开始测试
    print('------开始测试，按 ctrl+alt+s，可结束测试------')
    t1 = Thread(target=test_by_autogui, name='test_by_autogui') if g_mode == 'autogui' else Thread(
        target=test_by_javascript, name='test_by_javascript')
    t2 = Thread(target=record_data, name='record_data')
    t3 = Thread(target=keyboard_listener, name='keyboard_listener')
    g_start_time = int(datetime.now().timestamp())  # 记录开始时间戳
    g_is_running = True
    g_records = []
    t1.start()
    t2.start()
    t3.start()
