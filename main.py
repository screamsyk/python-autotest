# 内置模块
from configparser import ConfigParser
from threading import Thread
from time import sleep
from datetime import datetime
import json
import os

# 第三方模块
from selenium import webdriver
import keyboard
import pyautogui
import psutil
import openpyxl
import requests

# 全局数据
g_config = None
g_driver = None
g_start_time = None
g_end_time = None
g_is_running = False
g_records = []
g_pname = ''  # 浏览器进程名
g_pids = []  # 浏览器进程 id


def operate_map():
    '''(1)操作地图

    生成操作过程-->循环操作地图
    '''

    # 生成初始缩放过程
    init_scrolls = []
    init_scrolls.extend([-1000 for i in range(6)])
    init_scrolls.extend([1000 for i in range(12)])

    # 生成拖动过程（顺时针螺旋形）
    screen_width, screen_height = pyautogui.size()  # 屏幕大小
    center_x, center_y = screen_width/2, screen_height/2
    drags = []
    direction = 'right'  # 拖动方向
    num = 0  # 方向上的拖动次数
    total = g_config.getint('path_total')
    while total > 0:
        total -= 1
        if direction == 'right':  # 向右
            num = num+1
            drags.extend([{'direction': 'right', 'from': (center_x+300, center_y),
                           'to': (center_x-300, center_y)} for i in range(num)])
            direction = 'bottom'
        elif direction == 'bottom':  # 向下
            drags.extend([{'direction': 'bottom', 'from': (center_x, center_y+300),
                           'to': (center_x, center_y-300)} for i in range(num)])
            direction = 'left'
        elif direction == 'left':  # 向左
            num = num+1
            drags.extend([{'direction': 'left', 'from': (center_x-300, center_y),
                           'to': (center_x+300, center_y)} for i in range(num)])
            direction = 'top'
        elif direction == 'top':  # 向上
            drags.extend([{'direction': 'top', 'from': (center_x, center_y-300),
                           'to': (center_x, center_y+300)} for i in range(num)])
            direction = 'right'

    # 生成拖动时的缩放过程（缩-放-缩）
    scrolls = []
    scrolls.extend([-1000 for i in range(3)])
    scrolls.extend([1000 for i in range(4)])
    scrolls.extend([-1000 for i in range(1)])

    # 循环操作
    interval = g_config.getfloat('operate_interval')
    while not is_end():

        # 地图初始化
        print('地图初始化：', datetime.now().strftime('%H:%M:%S'))
        map_type = g_config['map_type']
        if map_type == 'd2c':
            script = f'var d2cMap=d2cMap||map||$map;d2cMap.setCenter([106.5590013579515, 29.55910442310595]);d2cMap.setZoom(12)'
        elif map_type == 'baidu':
            script = 'map.centerAndZoom(new BMap.Point(11863203.19134712,3426106.9054011325),13.7267941875)'
        elif map_type == 'gaode':
            script = 'themap.setZoomAndCenter(13,new AMap.LngLat(106.562411,29.556381))'
        g_driver.execute_script(script)
        sleep(interval)

        # 鼠标缩放地图
        for init_scroll in init_scrolls:
            if is_end():
                break
            pyautogui.moveTo(center_x, center_y)
            pyautogui.scroll(init_scroll)
            sleep(interval)

        # 鼠标拖动地图，并缩放地图
        drag_num = 0
        for drag in drags:
            if is_end():
                break
            drag_num += 1
            pyautogui.moveTo(drag['from'][0], drag['from'][1])
            pyautogui.dragTo(drag['to'][0], drag['to']
                             [1], duration=0.2)
            sleep(interval)
            # 每拖动 3 下，进行缩放
            if drag_num == 3:
                drag_num = 0
                for scroll in scrolls:
                    if is_end():
                        break
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.scroll(scroll)
                    sleep(interval)

    # 退出
    g_driver.close()
    g_driver.quit()


def record_data():
    '''(2)记录数据

    每秒统计下总的 CPU 使用率、内存使用率、内存使用大小，以及浏览器对应的数据
    '''
    interval = g_config.getfloat('record_interval')
    while g_is_running:
        browser_memory_percent = 0
        browser_memory = 0
        for pid in g_pids:
            try:
                process = psutil.Process(pid)
                browser_memory_percent += process.memory_percent(memtype='vms')
                browser_memory += process.memory_info().vms
            except:
                pass
        record = {
            'time': datetime.now().strftime('%H:%M:%S'),
            'cpu_percent': psutil.cpu_percent(),
            'memory_percent': psutil.virtual_memory().percent,
            'memory': round(psutil.virtual_memory().used/1024**2, 2),
            'browser_memory_percent': round(browser_memory_percent, 2),
            'browser_memory': round(browser_memory/1024**2, 2)
        }
        g_records.append(record)
        sleep(interval)
    stat_data()


def stat_data():
    '''(3)统计数据

    统计 CPU 平均使用率、内存平均使用率等
    '''
    xData = []
    yData_cpu = []
    yData_memory = []
    yData_browser_memory = []
    cpu_percent_sum = 0
    cpu_percent_avg = 0
    cpu_percent_over = 0
    memory_sum = 0
    memory_avg = 0
    memory_percent_sum = 0
    memory_percent_avg = 0
    memory_percent_over = 0
    browser_memory_percent_sum = 0
    browser_memory_percent_avg = 0
    browser_memory_sum = 0
    browser_memory_avg = 0
    for record in g_records:
        xData.append(record['time'])
        yData_cpu.append(record['cpu_percent'])
        yData_memory.append(record['memory_percent'])
        yData_browser_memory.append(record['browser_memory_percent'])
        cpu_percent_sum += record['cpu_percent']
        memory_sum += record['memory']
        memory_percent_sum += record['memory_percent']
        browser_memory_percent_sum += record['browser_memory_percent']
        browser_memory_sum += record['browser_memory']
        if record['cpu_percent'] >= 90:
            cpu_percent_over += 1
        if record['memory_percent'] >= 90:
            memory_percent_over += 1
    length = len(g_records)
    cpu_percent_avg = round(cpu_percent_sum/length, 2)
    memory_avg = round(memory_sum/length, 2)
    memory_percent_avg = round(memory_percent_sum/length, 2)
    browser_memory_percent_avg = round(browser_memory_percent_sum/length, 2)
    browser_memory_avg = round(browser_memory_sum/length, 2)
    stat_info = f'（1）CPU 平均使用率：{cpu_percent_avg}%；（2）CPU 使用率达 90% 及以上次数：{cpu_percent_over}；（3）内存平均使用率：{memory_percent_avg}%；（4）内存使用率达 90% 及以上次数：{memory_percent_over}；（5）内存平均使用大小：{memory_avg} MB；（6）内存平均使用率_浏览器：{browser_memory_percent_avg}%；（7）内存平均使用大小_浏览器：{browser_memory_avg} MB'
    save_to_excel(stat_info)
    save_to_json(stat_info)
    save_to_chart(xData, yData_cpu, yData_memory,
                  yData_browser_memory, stat_info)
    print('------程序停止------')
    print('测试结果：', stat_info)


def save_to_excel(stat_info):
    '''(4)保存数据到 excel
    '''
    filename = g_config['excel']
    if os.path.isfile(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.Workbook()
    ws = wb.active
    now_col = 1 if ws.max_column in [0, 1] else ws.max_column+2
    for i in range(0, 5):
        ws.column_dimensions[openpyxl.utils.get_column_letter(
            now_col+i)].width = 20
    name_cell = ws.cell(row=1, column=now_col, value=g_config['name'])
    name_cell.font = openpyxl.styles.Font(bold=True)
    ws.merge_cells(start_row=1, start_column=now_col,
                   end_row=1, end_column=now_col+4)
    ws.cell(
        row=2, column=now_col).value = f'统计（{datetime.fromtimestamp(g_start_time)}至{datetime.fromtimestamp(g_end_time)}）：{stat_info}'
    ws.merge_cells(start_row=2, start_column=now_col,
                   end_row=2, end_column=now_col+4)
    ws.cell(row=3, column=now_col).value = 'CUP 使用率（%）'
    ws.cell(row=3, column=now_col+1).value = '内存使用率（%）'
    ws.cell(row=3, column=now_col+2).value = '内存使用大小（MB）'
    ws.cell(row=3, column=now_col+3).value = '内存使用率_浏览器（%）'
    ws.cell(row=3, column=now_col+4).value = '内存使用大小_浏览器（MB）'
    now_row = 4
    for record in g_records:
        ws.cell(row=now_row, column=now_col).value = record['cpu_percent']
        ws.cell(row=now_row, column=now_col +
                1).value = record['memory_percent']
        ws.cell(row=now_row, column=now_col +
                2).value = record['memory']
        ws.cell(row=now_row, column=now_col +
                3).value = record['browser_memory_percent']
        ws.cell(row=now_row, column=now_col +
                4).value = record['browser_memory']
        now_row = now_row+1
    wb.save(filename=filename)


def save_to_json(stat_info):
    '''(5)保存数据到 json 文件
    '''
    name = g_config['name']
    obj = {'title': name, 'stat_info': stat_info, 'records': g_records}
    with open(f'{name}.json', 'w', encoding='utf-8') as f:
        f.write(json.dumps(obj, ensure_ascii=False))


def save_to_chart(xData, yData_cpu, yData_memory, yData_browser_memory, stat_info):
    '''(6)保存数据到图表
    '''
    name = g_config['name']
    with open('conf/chart_template.html', 'r', encoding='utf-8') as f:
        html = f.read()
        new_html = html.replace('{{title}}', name).replace(
            '{{text}}', name).replace('{{subtext}}', stat_info).replace('{{dataScript}}', f'''
    <script>
        var xData={json.dumps(xData)};
        var yData_cpu={json.dumps(yData_cpu)};
        var yData_memory={json.dumps(yData_memory)};
        var yData_browser_memory={json.dumps(yData_browser_memory)};
    </script>
        ''')
    with open(f'{name}.html', 'w', encoding='utf-8') as new_f:
        new_f.write(new_html)


def start():
    '''(7)开始测试
    '''
    global g_start_time
    global g_is_running
    global g_pids
    keyboard.remove_hotkey('ctrl+alt+s')
    print('------开始测试，通过快捷键 ctrl+alt+e 结束测试------')
    keyboard.add_hotkey('ctrl+alt+e', end)
    g_start_time = int(datetime.now().timestamp())
    g_is_running = True
    g_pids = [process.info['pid'] for process in psutil.process_iter(
        attrs=['pid', 'name']) if process.info['name'] == g_pname]
    t1 = Thread(target=operate_map, name='operate_map')
    t2 = Thread(target=record_data, name='record_data')
    t1.start()
    t2.start()


def end():
    '''(8)结束测试
    '''
    global g_is_running
    g_is_running = False


def is_end():
    '''(9)是否结束测试了
    '''
    global g_end_time
    global g_is_running
    g_end_time = int(datetime.now().timestamp())
    time = g_config.getfloat('minute')*60
    if g_end_time-g_start_time > time:
        g_is_running = False
    return not g_is_running


def main():
    global g_config
    global g_driver
    global g_pname

    # 读取配置信息
    config = ConfigParser()
    config.read('conf/config.ini', encoding="utf-8-sig")
    if not config.has_section('config'):
        print('------无正确配置信息，程序结束------')
        return
    g_config = config['config'] or {}

    # 驱动浏览器
    browser = g_config['browser'] or 'chrome'
    if browser == 'chrome':
        path = 'webdriver/chromedriver'
        g_driver = webdriver.Chrome(executable_path=path)
        g_pname = 'chrome.exe'
    elif browser == 'firefox':
        path = 'webdriver/geckodriver'
        g_driver = webdriver.Firefox(executable_path=path)
        g_pname = 'firefox.exe'
    elif browser == 'edge':
        path = 'webdriver/MicrosoftWebDriver.exe'
        g_driver = webdriver.Edge(executable_path=path)
        g_pname = 'MicrosoftEdge.exe'
    elif browser == 'ie':
        path = 'webdriver/IEDriverServer.exe'
        g_driver = webdriver.Ie(executable_path=path)
        g_pname = 'iexplore.exe'
    g_driver.maximize_window()
    g_driver.get(g_config['url'])

    # 是否需要登录
    if g_config.getboolean('need_login'):
        sleep(1)
        try:
            inputs = g_driver.find_elements_by_tag_name('input')
            inputs[0].send_keys(g_config['username'])
            inputs[1].send_keys(g_config['password'])
        except:
            pass

    # 通过快捷键开始测试
    print('------通过快捷键 ctrl+alt+s 开始测试------')
    keyboard.add_hotkey('ctrl+alt+s', start)
    keyboard.wait()


if __name__ == '__main__':
    main()
